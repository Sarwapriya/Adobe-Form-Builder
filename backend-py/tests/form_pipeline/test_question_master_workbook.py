"""Python port of packages/shared/tests/excel/questionMasterWorkbook.test.ts."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.form_pipeline.excel.question_master_rows import QuestionMasterRow
from app.form_pipeline.excel.question_master_workbook import build_question_master_workbook

_SAMPLE_ROWS: list[QuestionMasterRow] = [
    QuestionMasterRow(
        division="MX",
        project="F2H26",
        subsidiary="SEIL",
        country_alpha_2="PS",
        locale="ar_PS",
        question_code="FIRSTNAME",
        question_text_full="الاسم الأول",
        question_text_alias="firstname",
        mandatory_yn="Y",
        local_yn="Standard",
        type="Free text",
        answer_code="firstName",
        answer_text_full="الاسم الأول",
        answer_text_alias="firstname",
    ),
    QuestionMasterRow(
        division="MX",
        project="F2H26",
        subsidiary="SEIL",
        country_alpha_2="PS",
        locale="ar_PS",
        question_code="HPP_CODE",
        question_text_full="رقم الهاتف المتحرّك",
        question_text_alias="mobileNumber",
        mandatory_yn="N",
        local_yn="Local",
        type="Free text",
        answer_code="mobileNumber",
        answer_text_full="رقم الهاتف المتحرّك",
        answer_text_alias="mobileNumber",
    ),
]

_COLUMNS = (
    "division",
    "project",
    "subsidiary",
    "country_alpha_2",
    "locale",
    "question_code",
    "question_text_full",
    "question_text_alias",
    "mandatory_yn",
    "local_yn",
    "type",
    "answer_code",
    "answer_text_full",
    "answer_text_alias",
)


def test_round_trips_headers_and_row_content():
    raw = build_question_master_workbook(_SAMPLE_ROWS)
    workbook = load_workbook(BytesIO(raw))

    assert workbook.sheetnames == ["Question Master"]
    sheet = workbook["Question Master"]
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == _COLUMNS
    assert len(rows) - 1 == 2  # header + 2 data rows

    row0 = dict(zip(_COLUMNS, rows[1]))
    assert row0["division"] == "MX"
    assert row0["project"] == "F2H26"
    assert row0["subsidiary"] == "SEIL"
    assert row0["country_alpha_2"] == "PS"
    assert row0["locale"] == "ar_PS"
    assert row0["question_code"] == "FIRSTNAME"
    assert row0["question_text_full"] == "الاسم الأول"
    assert row0["question_text_alias"] == "firstname"
    assert row0["mandatory_yn"] == "Y"
    assert row0["local_yn"] == "Standard"
    assert row0["type"] == "Free text"
    assert row0["answer_code"] == "firstName"

    row1 = dict(zip(_COLUMNS, rows[2]))
    assert row1["question_code"] == "HPP_CODE"
    assert row1["local_yn"] == "Local"
    assert row1["answer_code"] == "mobileNumber"


def test_produces_empty_but_valid_sheet_for_zero_rows():
    raw = build_question_master_workbook([])
    workbook = load_workbook(BytesIO(raw))
    sheet = workbook["Question Master"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows == [_COLUMNS]


def test_returns_real_bytes_that_can_be_written_and_read_back(tmp_path):
    raw = build_question_master_workbook(_SAMPLE_ROWS)
    assert isinstance(raw, bytes)

    file_path = tmp_path / "question-master.xlsx"
    file_path.write_bytes(raw)

    workbook = load_workbook(file_path)
    assert workbook.sheetnames == ["Question Master"]
